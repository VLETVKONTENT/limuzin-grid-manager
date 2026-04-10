from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from limuzin_grid_manager.core.crs import ck42_to_wgs84, infer_gk_zone, make_transformer_for_zone
from limuzin_grid_manager.core.points import PointRecord, normalize_point_date, parse_point_coordinates

_EXPECTED_HEADERS = ("фио", "дата", "координаты")


@dataclass(frozen=True)
class PointImportError:
    source_row: int
    message: str

    def __post_init__(self) -> None:
        object.__setattr__(self, "source_row", int(self.source_row))
        object.__setattr__(self, "message", str(self.message).strip())


@dataclass(frozen=True)
class PointImportResult:
    sheet_name: str
    records: tuple[PointRecord, ...]
    errors: tuple[PointImportError, ...]
    total_rows: int

    def __post_init__(self) -> None:
        object.__setattr__(self, "sheet_name", str(self.sheet_name).strip())
        object.__setattr__(self, "records", tuple(self.records))
        object.__setattr__(self, "errors", tuple(self.errors))
        object.__setattr__(self, "total_rows", max(0, int(self.total_rows)))

    @property
    def is_exportable(self) -> bool:
        return bool(self.records) and not self.errors

    @property
    def summary(self) -> str:
        lines = [
            f"Лист: {self.sheet_name or 'без имени'}",
            f"Строк данных: {self.total_rows}.",
            f"Корректных точек: {len(self.records)}.",
        ]
        if self.errors:
            lines.append(f"Ошибок: {len(self.errors)}. Экспорт заблокирован.")
        else:
            lines.append("Ошибок нет. Экспорт доступен.")
        return "\n".join(lines)


def import_points_from_excel(path: Path) -> PointImportResult:
    workbook_path = Path(path)
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("Импорт точек поддерживает только файлы .xlsx.")
    if not workbook_path.exists() or not workbook_path.is_file():
        raise ValueError(f"Excel-файл не найден: {workbook_path}.")

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Не удалось открыть Excel-файл: {exc}") from exc

    try:
        worksheet = _find_first_sheet_with_data(workbook.worksheets)
        return _import_points_from_sheet(worksheet)
    finally:
        workbook.close()


def _find_first_sheet_with_data(worksheets: list[Worksheet]) -> Worksheet:
    for worksheet in worksheets:
        for row in worksheet.iter_rows(values_only=True):
            if _row_has_data(row):
                return worksheet
    raise ValueError("В workbook нет ни одного листа с данными.")


def _import_points_from_sheet(worksheet: Worksheet) -> PointImportResult:
    records: list[PointRecord] = []
    errors: list[PointImportError] = []
    total_rows = 0
    header_indexes: tuple[int, int, int] | None = None

    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if not _row_has_data(row):
            continue

        if header_indexes is None:
            header_indexes = _parse_header(row, sheet_name=worksheet.title)
            continue

        total_rows += 1
        shape_error = _validate_row_shape(row, header_indexes)
        if shape_error is not None:
            errors.append(PointImportError(source_row=row_number, message=shape_error))
            continue

        name_value, date_value, coordinates_value = _extract_cells(row, header_indexes)
        try:
            record = _build_record(
                row_number=row_number,
                name_value=name_value,
                date_value=date_value,
                coordinates_value=coordinates_value,
            )
        except ValueError as exc:
            errors.append(PointImportError(source_row=row_number, message=str(exc)))
            continue
        records.append(record)

    if header_indexes is None:
        raise ValueError(f"Лист '{worksheet.title}' не содержит непустого заголовка.")

    return PointImportResult(
        sheet_name=worksheet.title,
        records=tuple(records),
        errors=tuple(errors),
        total_rows=total_rows,
    )


def _parse_header(row: tuple[object, ...], *, sheet_name: str) -> tuple[int, int, int]:
    non_empty_headers = [
        (index, _normalize_header_cell(value))
        for index, value in enumerate(row)
        if not _is_empty_cell(value)
    ]
    header_names = tuple(name for _, name in non_empty_headers)
    if header_names != _EXPECTED_HEADERS:
        expected = " | ".join(header.title() for header in _EXPECTED_HEADERS)
        actual = " | ".join(header_names) or "пусто"
        raise ValueError(
            f"Лист '{sheet_name}' должен начинаться с заголовка '{expected}'. Получено: '{actual}'."
        )
    return tuple(index for index, _ in non_empty_headers)  # type: ignore[return-value]


def _validate_row_shape(row: tuple[object, ...], header_indexes: tuple[int, int, int]) -> str | None:
    extra_cells = [
        value
        for index, value in enumerate(row)
        if index not in header_indexes and not _is_empty_cell(value)
    ]
    if extra_cells:
        return "Строка должна содержать только колонки 'ФИО', 'Дата' и 'Координаты'."
    return None


def _extract_cells(
    row: tuple[object, ...],
    header_indexes: tuple[int, int, int],
) -> tuple[object, object, object]:
    values = [row[index] if index < len(row) else None for index in header_indexes]
    return values[0], values[1], values[2]


def _build_record(
    *,
    row_number: int,
    name_value: object,
    date_value: object,
    coordinates_value: object,
) -> PointRecord:
    if _is_empty_cell(name_value):
        raise ValueError("Поле 'ФИО' не заполнено.")
    name = str(name_value).strip()
    if _is_empty_cell(date_value):
        raise ValueError("Поле 'Дата' не заполнено.")
    if _is_empty_cell(coordinates_value):
        raise ValueError("Поле 'Координаты' не заполнено.")

    display_date = normalize_point_date(date_value)
    x, y = parse_point_coordinates(str(coordinates_value))
    zone = infer_gk_zone(y)
    transformer = make_transformer_for_zone(zone)
    lon, lat = ck42_to_wgs84(x, y, transformer)

    return PointRecord(
        name=name,
        source_date=_stringify_source_value(date_value),
        display_date=display_date,
        x=x,
        y=y,
        zone=zone,
        lon=lon,
        lat=lat,
        source_row=row_number,
    )


def _normalize_header_cell(value: object) -> str:
    return " ".join(str(value).strip().split()).casefold()


def _row_has_data(row: tuple[object, ...]) -> bool:
    return any(not _is_empty_cell(value) for value in row)


def _is_empty_cell(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _stringify_source_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
